"""
Generación de reportes para Innova Corp.
- PDF con reportlab (reporte general, reporte semanal, facturas)
- Excel con openpyxl (inventario)
- CSV con la librería estándar (técnicos)

Todo se genera en memoria (io.BytesIO / io.StringIO) y se sirve directo
como descarga, sin dejar archivos temporales en el servidor.
"""
import csv
import io
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ============================================================
# Helpers de estilo para PDFs
# ============================================================
def _estilos():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='TituloInnova', fontSize=18, textColor=colors.HexColor('#1b1f2a'),
        spaceAfter=4, fontName='Helvetica-Bold'
    ))
    styles.add(ParagraphStyle(
        name='SubtituloInnova', fontSize=10, textColor=colors.HexColor('#6b7280'),
        spaceAfter=14
    ))
    styles.add(ParagraphStyle(
        name='SeccionInnova', fontSize=13, textColor=colors.HexColor('#667eea'),
        spaceBefore=14, spaceAfter=8, fontName='Helvetica-Bold'
    ))
    return styles


def _tabla_estilo(header_bg='#667eea'):
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_bg)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e2e2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7fb')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])


# ============================================================
# a) Reporte General del Sistema (PDF)
# ============================================================
def generar_reporte_general_pdf(servicios, clientes):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = _estilos()
    elementos = []

    elementos.append(Paragraph('Innova Corp — Reporte General del Sistema', styles['TituloInnova']))
    elementos.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['SubtituloInnova']))

    valor_inventario = sum(float(s.precio_unitario) * s.stock for s in servicios)
    resumen_data = [
        ['Indicador', 'Valor'],
        ['Total de servicios/productos', str(len(servicios))],
        ['Valor total de inventario', f"${valor_inventario:,.2f}"],
        ['Total de clientes registrados', str(len(clientes))],
    ]
    t = Table(resumen_data, colWidths=[9 * cm, 6 * cm])
    t.setStyle(_tabla_estilo())
    elementos.append(t)

    elementos.append(Paragraph('Directorio de Clientes', styles['SeccionInnova']))
    if clientes:
        datos_clientes = [['Nombre / Razón social', 'RIF', 'Dirección fiscal']]
        for c in clientes:
            datos_clientes.append([c.nombre_razon_social, c.rif or '-', c.direccion_fiscal or '-'])
        tc = Table(datos_clientes, colWidths=[6 * cm, 3.5 * cm, 6.5 * cm])
        tc.setStyle(_tabla_estilo())
        elementos.append(tc)
    else:
        elementos.append(Paragraph('No hay clientes registrados.', styles['Normal']))

    doc.build(elementos)
    buffer.seek(0)
    return buffer


# ============================================================
# b) Reporte Semanal de Pedidos (PDF)
# ============================================================
def generar_reporte_semanal_pdf(pedidos):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = _estilos()
    elementos = []

    desde = datetime.utcnow() - timedelta(days=7)
    elementos.append(Paragraph('Innova Corp — Reporte Semanal de Pedidos', styles['TituloInnova']))
    elementos.append(Paragraph(
        f"Del {desde.strftime('%d/%m/%Y')} al {datetime.utcnow().strftime('%d/%m/%Y')}",
        styles['SubtituloInnova']
    ))

    if pedidos:
        data = [['#', 'Cliente', 'Fecha', 'Estado', 'Total']]
        total_periodo = 0
        for p in pedidos:
            data.append([
                str(p.id_pedido),
                p.cliente.nombre_razon_social if p.cliente else '-',
                p.fecha_pedido.strftime('%d/%m/%Y') if p.fecha_pedido else '-',
                p.estado.replace('_', ' '),
                f"${float(p.total):,.2f}"
            ])
            total_periodo += float(p.total)
        t = Table(data, colWidths=[1.5 * cm, 6 * cm, 3 * cm, 3 * cm, 3 * cm])
        t.setStyle(_tabla_estilo())
        elementos.append(t)
        elementos.append(Spacer(1, 12))
        elementos.append(Paragraph(f"<b>Total del período: ${total_periodo:,.2f}</b>", styles['Normal']))
    else:
        elementos.append(Paragraph('No se registraron pedidos en los últimos 7 días.', styles['Normal']))

    doc.build(elementos)
    buffer.seek(0)
    return buffer


# ============================================================
# c) Listado de Inventario (Excel .xlsx)
# ============================================================
def generar_inventario_xlsx(servicios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    encabezados = ['ID', 'Nombre', 'Descripción', 'Precio', 'Stock', 'Valor Total']
    ws.append(encabezados)

    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = header_fill

    for s in servicios:
        valor_total = float(s.precio_unitario) * s.stock
        ws.append([
            s.id_servicio, s.nombre, s.descripcion or '-',
            float(s.precio_unitario), s.stock, valor_total
        ])

    anchos = [6, 28, 40, 12, 10, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# d) Técnicos Disponibles (CSV)
# ============================================================
def generar_tecnicos_csv(tecnicos):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Nombre Completo', 'Cédula', 'Especialidad', 'Teléfono'])
    for t in tecnicos:
        writer.writerow([t.nombre_completo, t.cedula, t.especialidad or '-', t.telefono or '-'])
    buffer.seek(0)
    return buffer


# ============================================================
# e) Factura individual (PDF) — con IVA 16% y numeración fiscal
# ============================================================
def generar_factura_pdf(pedido, factura, config):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = _estilos()
    elementos = []

    elementos.append(Paragraph(config.EMPRESA_RAZON_SOCIAL, styles['TituloInnova']))
    elementos.append(Paragraph(
        f"RIF: {config.EMPRESA_RIF} · {config.EMPRESA_DIRECCION_FISCAL}",
        styles['SubtituloInnova']
    ))

    datos_factura = [
        ['Número de Factura', factura.numero_factura],
        ['Número de Control', factura.numero_control],
        ['Fecha de emisión', factura.fecha_emision.strftime('%d/%m/%Y')],
        ['Cliente', pedido.cliente.nombre_razon_social],
        ['RIF Cliente', factura.rif_cliente_snapshot],
        ['Dirección fiscal', pedido.cliente.direccion_fiscal],
    ]
    tf = Table(datos_factura, colWidths=[5 * cm, 10 * cm])
    tf.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tf)
    elementos.append(Spacer(1, 14))

    elementos.append(Paragraph('Detalle', styles['SeccionInnova']))
    detalle_data = [['Descripción', 'Cantidad', 'Precio Unit.', 'Subtotal']]
    for d in pedido.detalles:
        detalle_data.append([
            d.servicio.nombre if d.servicio else '-',
            str(d.cantidad),
            f"${float(d.precio_unitario):,.2f}",
            f"${float(d.subtotal_linea):,.2f}"
        ])
    td = Table(detalle_data, colWidths=[7 * cm, 2.5 * cm, 3 * cm, 3 * cm])
    td.setStyle(_tabla_estilo())
    elementos.append(td)
    elementos.append(Spacer(1, 14))

    totales_data = [
        ['Base imponible', f"${float(factura.base_imponible):,.2f}"],
        [f"IVA ({float(factura.porcentaje_iva):.0f}%)", f"${float(factura.monto_iva):,.2f}"],
        ['TOTAL A PAGAR', f"${float(factura.total):,.2f}"],
    ]
    tt = Table(totales_data, colWidths=[10 * cm, 5 * cm])
    tt.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#667eea')),
        ('TOPPADDING', (0, -1), (-1, -1), 8),
    ]))
    elementos.append(tt)

    doc.build(elementos)
    buffer.seek(0)
    return buffer
